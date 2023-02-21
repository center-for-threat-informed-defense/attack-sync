import json
import re
from pathlib import Path
from typing import Dict

from loguru import logger
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import get_column_letter


def get_attack_id(changelog_attack_object: dict):
    attack_id = None
    external_references = changelog_attack_object.get("external_references")

    if external_references:
        attack_source = external_references[0]
        if attack_source.get("external_id") and attack_source.get("source_name") == "mitre-attack":
            attack_id = attack_source["external_id"]

    return attack_id


def get_changed_techniques(changelog: Dict[str, dict]):
    logger.debug("Getting techniques that have changes between ATT&CK versions")
    changed_techniques = {}
    for domain, change_info in changelog.items():
        if domain == "new-contributors":
            continue

        technique_changes = change_info["techniques"]
        for change_type, changes in technique_changes.items():
            for technique_object in changes:
                attack_id = get_attack_id(technique_object)
                technique_object["CHANGE_TYPE"] = change_type
                technique_object["DOMAIN"] = domain
                changed_techniques[attack_id] = technique_object

    return changed_techniques


def main():
    old_version = "10.1"
    new_version = "12.1"
    changelog_json_file = f"data/attack-changelog-v{old_version}-v{new_version}.json"
    original_mapping_file = "data/nist800-53-r4-mappings.xlsx"
    output_dir = "output"
    final_workbook = f"{output_dir}/mapping_diff_attack_v{old_version}-v{new_version}.xlsx"

    with open(changelog_json_file, "r") as file:
        changelog = json.load(file)

    changed_techniques = get_changed_techniques(changelog=changelog)

    logger.debug(f"Loading: {original_mapping_file}")
    mapping_workbook = load_workbook(filename=original_mapping_file)
    mapping_sheet = mapping_workbook["Sheet1"]
    initial_cols = 5
    cols_per_change = 3
    num_changes = 5 # there are 5 hard coded checks for changes below
    added_cols = cols_per_change * num_changes
    total_cols = initial_cols + added_cols

    def write_change(change_type, old_value, new_value):
        nonlocal start_col
        row[start_col].value = change_type
        row[start_col+1].value = old_value
        row[start_col+2].value = new_value
        start_col += 3

    logger.debug("Providing context for ATT&CK Techniques that have changed")
    for row in mapping_sheet.iter_rows(max_col=total_cols):
        control_id = row[0].value
        technique_id = row[3].value
        start_col = 5

        # The first row needs column headers
        if control_id == "Control ID":
            for _ in range(5):
                write_change("Change", "Old Value", "New Value")
            continue

        if technique_id in changed_techniques:
            changed_technique = changed_techniques[technique_id]
            change_type = changed_technique["CHANGE_TYPE"]

            if change_type == "additions":
                continue

            detailed_diff = json.loads(changed_technique.get("detailed_diff", "{}"))
            if detailed_diff:
                regex = r"^root\['(?P<top_stix_key>[^\']*)'\](?P<the_rest>.*)$"
                for detailed_change_type, detailed_changes in detailed_diff.items():
                    if detailed_change_type == "values_changed":
                        for detailed_change, values in detailed_changes.items():
                            matches = re.search(regex, detailed_change)
                            top_stix_key = matches.group("top_stix_key")
                            the_rest = matches.group("the_rest")
                            stix_field = f"{top_stix_key}{the_rest}"

                            old_value = values["old_value"]
                            new_value = values["new_value"]

                            if stix_field == "description":
                                write_change(
                                    "Modified Description",
                                    old_value,
                                    new_value,
                                )

            changelog_mitigations = changed_technique.get("changelog_mitigations")
            if changelog_mitigations:
                if changelog_mitigations["new"]:
                    new_value = ""
                    for new_mitigation in changelog_mitigations["new"]:
                        new_value += f"{new_mitigation}\n"
                    write_change(
                        "New Mitigations",
                        "",
                        new_value.strip(),
                    )
                if changelog_mitigations["dropped"]:
                    new_value = ""
                    for dropped_mitigation in changelog_mitigations["dropped"]:
                        new_value += f"{dropped_mitigation}\n"
                    write_change(
                        "Dropped Mitigations",
                        "",
                        new_value.strip(),
                    )
            changelog_detections = changed_technique.get("changelog_detections")
            if changelog_detections:
                if changelog_detections["new"]:
                    new_value = ""
                    for new_detection in changelog_detections["new"]:
                        new_value += f"{new_detection}\n"
                    write_change(
                        "New Detections",
                        "",
                        new_value.strip(),
                    )
                if changelog_detections["dropped"]:
                    new_value = ""
                    for dropped_detections in changelog_detections["dropped"]:
                        new_value += f"{dropped_detections}\n"
                    write_change(
                        "Dropped Detections",
                        "",
                        new_value.strip(),
                    )

    ###########
    # Style it!
    ###########
    logger.debug("Widening new columns...")
    for i in range(0, num_changes):
        col_idx = initial_cols + 1 + i * cols_per_change
        print("set ", get_column_letter(col_idx), "to 18")
        mapping_sheet.column_dimensions[get_column_letter(col_idx)].width = 18
        for j in range(1, cols_per_change):
            print("set ", get_column_letter(col_idx+j), "to 100")
            mapping_sheet.column_dimensions[get_column_letter(col_idx + j)].width = 100

    # Wrap text
    logger.debug("Wrapping text for new columns...")
    for row in mapping_sheet.rows:
        for col_idx in range(initial_cols + 1, total_cols):
            row[col_idx].alignment = Alignment(wrap_text=True)

    logger.debug(f"Creating output directory: {output_dir}")
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    logger.debug(f"Saving file: {final_workbook}")
    mapping_workbook.save(filename=final_workbook)


if __name__ == "__main__":
    main()
