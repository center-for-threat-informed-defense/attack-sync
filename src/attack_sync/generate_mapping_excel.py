import json
import re
import string
from argparse import ArgumentParser, BooleanOptionalAction
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict

from loguru import logger
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock
from openpyxl.cell.text import InlineFont
from openpyxl.styles import Alignment, Font
from openpyxl.utils.cell import get_column_letter

TID_RE = re.compile(r"^T\d\d\d\d(\.\d\d\d)?$")


def get_attack_id(changelog_attack_object: dict):
    attack_id = None
    external_references = changelog_attack_object.get("external_references")

    if external_references:
        attack_source = external_references[0]
        if (
            attack_source.get("external_id")
            and attack_source.get("source_name") == "mitre-attack"
        ):
            attack_id = attack_source["external_id"]

    return attack_id


def get_changed_techniques(changelog: Dict[str, dict]):
    logger.info("Getting techniques that have changes between ATT&CK versions")
    changed_techniques = {}
    for domain, change_info in changelog.items():
        if domain == "new-contributors":
            continue

        technique_changes = change_info["techniques"]
        for change_type, changes in technique_changes.items():
            for technique_object in changes:
                attack_id = get_attack_id(technique_object)
                technique_object["CHANGE_TYPE"] = change_type
                technique_object["DOMAIN"] = domain
                changed_techniques[attack_id] = technique_object

    return changed_techniques


def parse_args():
    parser = ArgumentParser(
        description="Annotate a mappings spreadsheet with notes from an ATT&CK changelog."
    )
    parser.add_argument(
        "changelog",
        type=str,
        help="Path to ATT&CK changelog",
    )
    parser.add_argument(
        "mapping_in",
        type=str,
        help="Path to input mapping spreadsheet (xlsx).",
    )
    parser.add_argument(
        "attack_id_column",
        type=str,
        help="The column letter in the spreadsheet that contains ATT&CK object IDs.",
    )
    parser.add_argument(
        "attack_name_column",
        type=str,
        help="The column letter in the spreadsheet that contains ATT&CK object names.",
    )
    parser.add_argument(
        "output_column",
        type=str,
        help="The column letter where ATT&CK sync output should begin.",
    )
    parser.add_argument(
        "mapping_out",
        type=str,
        help="Path to save output mapping spreadsheet (xlsx).",
    )
    parser.add_argument(
        "--column-headers",
        action=BooleanOptionalAction,
        default=True,
        help="Indicates whether the source spreadsheet has column headers or not.",
    )
    return parser.parse_args()


def column_letter_to_number(letter):
    return string.ascii_uppercase.index(letter)


def main():
    args = parse_args()

    changelog_json_file = Path(args.changelog)
    original_mapping_file = Path(args.mapping_in)
    final_workbook = Path(args.mapping_out)

    with changelog_json_file.open("r") as file:
        changelog = json.load(file)

    changed_techniques = get_changed_techniques(changelog=changelog)

    logger.info(f"Loading: {original_mapping_file}")
    mapping_workbook = load_workbook(filename=original_mapping_file)
    mapping_sheet = mapping_workbook.worksheets[0]
    start_col = column_letter_to_number(args.output_column)
    attack_id_col = column_letter_to_number(args.attack_id_column)
    attack_name_col = column_letter_to_number(args.attack_name_column)
    cols_per_change = 2  # each change has two columns: change type and change value
    num_changes = 5  # there are 5 hard coded checks for changes below
    added_cols = cols_per_change * num_changes
    total_cols = start_col + added_cols

    add_style = InlineFont(b=True, color="339933")
    del_style = InlineFont(strike=True, color="cc0000")
    strike_font = Font(strike=True, color="cc0000")

    def get_tid_from_ext_refs(external_refs):
        for external_ref in external_refs:
            external_id = external_ref.get("external_id", "")
            if TID_RE.match(external_id):
                return external_id
        raise Exception(f"No technique ID found in externals refs: {external_refs}")

    def write_change(row, column, change_type, change_value):
        row[column].value = change_type
        row[column + 1].value = change_value
        # Return the number of columns written -- hardcoded at 2 for now
        return 2

    logger.info("Providing context for ATT&CK Techniques that have changed")

    row_iterator = mapping_sheet.iter_rows(max_col=total_cols)

    # If there are column headers, then add new headers for the ATT&CK Sync columns.
    if args.column_headers:
        row = next(row_iterator)
        for column in range(start_col, total_cols, cols_per_change):
            write_change(row, column, "Change Type", "Change Value")

    for row in row_iterator:
        technique_id = row[attack_id_col].value
        current_col = start_col

        if technique_id in changed_techniques:
            changed_technique = changed_techniques[technique_id]
            change_type = changed_technique["CHANGE_TYPE"]

            if change_type == "additions":
                continue

            revoked = changed_technique.get("revoked")
            if revoked:
                revoked_by = changed_technique.get("revoked_by")
                if revoked_by:
                    revoked_by_tid = get_tid_from_ext_refs(
                        revoked_by["external_references"]
                    )
                    revocation = f'Superseded by "{revoked_by_tid} {revoked_by["name"]}" - {revoked_by["description"]}'
                else:
                    revocation = "No superseding technique."
                current_col += write_change(row, current_col, "Revoked By", revocation)
                # Strike out the technique ID/name columns
                row[attack_id_col].font = strike_font
                row[attack_name_col].font = strike_font

            detailed_diff = json.loads(changed_technique.get("detailed_diff", "{}"))
            if detailed_diff:
                regex = r"^root\['(?P<top_stix_key>[^\']*)'\](?P<the_rest>.*)$"
                for detailed_change_type, detailed_changes in detailed_diff.items():
                    if detailed_change_type == "values_changed":
                        for detailed_change, values in detailed_changes.items():
                            matches = re.search(regex, detailed_change)
                            top_stix_key = matches.group("top_stix_key")
                            the_rest = matches.group("the_rest")
                            stix_field = f"{top_stix_key}{the_rest}"

                            old_text = str(values["old_value"]).split(" ")
                            new_text = str(values["new_value"]).split(" ")
                            cell_contents = list()
                            diff = SequenceMatcher(
                                None, old_text, new_text, autojunk=False
                            )

                            for (
                                opcode,
                                old_start,
                                old_end,
                                new_start,
                                new_end,
                            ) in diff.get_opcodes():
                                old = " " + " ".join(old_text[old_start:old_end]) + " "
                                new = " " + " ".join(new_text[new_start:new_end]) + " "
                                match opcode:
                                    case "replace":
                                        cell_contents.append(TextBlock(del_style, old))
                                        cell_contents.append(TextBlock(add_style, new))
                                    case "delete":
                                        cell_contents.append(TextBlock(del_style, old))
                                    case "insert":
                                        cell_contents.append(TextBlock(add_style, new))
                                    case "equal":
                                        cell_contents.append(new)

                            if stix_field == "description":
                                current_col += write_change(
                                    row,
                                    current_col,
                                    "Modified Description",
                                    CellRichText(cell_contents),
                                )

            changelog_mitigations = changed_technique.get("changelog_mitigations")
            if changelog_mitigations:
                if changelog_mitigations["new"]:
                    new_value = ""
                    for new_mitigation in changelog_mitigations["new"]:
                        new_value += f"{new_mitigation}\n"
                    current_col += write_change(
                        row,
                        current_col,
                        "New Mitigations",
                        CellRichText([TextBlock(add_style, new_value.strip())]),
                    )
                if changelog_mitigations["dropped"]:
                    new_value = ""
                    for dropped_mitigation in changelog_mitigations["dropped"]:
                        new_value += f"{dropped_mitigation}\n"
                    current_col += write_change(
                        row,
                        current_col,
                        "Dropped Mitigations",
                        CellRichText([TextBlock(del_style, new_value.strip())]),
                    )
            changelog_detections = changed_technique.get("changelog_detections")
            if changelog_detections:
                if changelog_detections["new"]:
                    new_value = ""
                    for new_detection in changelog_detections["new"]:
                        new_value += f"{new_detection}\n"
                    current_col += write_change(
                        row,
                        current_col,
                        "New Detections",
                        CellRichText([TextBlock(add_style, new_value.strip())]),
                    )
                if changelog_detections["dropped"]:
                    new_value = ""
                    for dropped_detections in changelog_detections["dropped"]:
                        new_value += f"{dropped_detections}\n"
                    current_col += write_change(
                        row,
                        current_col,
                        "Dropped Detections",
                        CellRichText([TextBlock(del_style, new_value.strip())]),
                    )

    additions = {
        tid: technique
        for (tid, technique) in changed_techniques.items()
        if technique["CHANGE_TYPE"] == "additions"
    }
    for tid, addition in sorted(additions.items()):
        new_row = [""] * total_cols
        new_row[attack_id_col] = tid
        new_row[attack_name_col] = addition["name"]
        new_row[start_col] = "New Technique"
        new_row[start_col + 1] = CellRichText(
            [TextBlock(add_style, addition["description"])]
        )
        mapping_sheet.append(new_row)

    logger.info("Widening new columns...")
    for col_idx in range(start_col + 1, total_cols + 1, cols_per_change):
        mapping_sheet.column_dimensions[get_column_letter(col_idx)].width = 18
        mapping_sheet.column_dimensions[get_column_letter(col_idx + 1)].width = 100

    logger.info("Wrapping text for new columns...")
    for row in mapping_sheet.rows:
        for col_idx in range(start_col, total_cols):
            row[col_idx].alignment = Alignment(wrap_text=True)

    logger.info(f"Saving file: {final_workbook}")
    final_workbook.parent.mkdir(parents=True, exist_ok=True)
    mapping_workbook.save(filename=final_workbook)


if __name__ == "__main__":
    main()
